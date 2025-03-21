from aiogram import Bot, Dispatcher, types
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.filters import Command
from datetime import datetime
import openpyxl
import pytz
import os
import asyncio
from telethon import TelegramClient, events
from telethon.tl.types import PeerChannel
from config import api_hash, api_id, BOT_TOKEN

# Инициализация бота и диспетчера
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

# Инициализация клиента Telethon
client = TelegramClient('session_name', api_id, api_hash)

# Путь к Excel-файлу
EXCEL_FILE = 'channel_posts.xlsx'

# Создаем Excel-файл, если он не существует
if not os.path.exists(EXCEL_FILE):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Posts"
    sheet.append(["Дата публикации", "Название канала", "Содержание поста", "Количество реакций"])
    workbook.save(EXCEL_FILE)

# Функция для добавления данных в Excel
def add_to_excel(date, channel_name, content, reactions):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([date, channel_name, content, reactions])
    workbook.save(EXCEL_FILE)

# Обработчик новых постов в каналах
@client.on(events.NewMessage)
async def new_post_listener(event):
    # Проверяем, что сообщение пришло из канала
    if isinstance(event.peer_id, PeerChannel):
        try:
            # Получаем информацию о канале
            channel = await client.get_entity(event.peer_id)
            channel_name = channel.title

            # Проверяем, что это пост, а не комментарий
            if event.is_group or event.is_channel:
                # Получаем время публикации в UTC
                post_date_utc = event.message.date

                # Преобразуем время в московское
                moscow_tz = pytz.timezone('Europe/Moscow')
                post_date_moscow = post_date_utc.astimezone(moscow_tz)

                # Форматируем дату (без времени)
                post_date = post_date_moscow.strftime('%Y-%m-%d')  # Только дата

                # Получаем содержание поста
                post_content = event.message.text or "Медиа-сообщение (без текста)"
                reactions = event.message.reactions.count if event.message.reactions else 0

                # Добавляем данные в Excel
                add_to_excel(post_date, channel_name, post_content, reactions)
                print(f"Добавлен пост из канала {channel_name} в {post_date} (Московское время)")
        except Exception as e:
            print(f"Ошибка при обработке поста: {e}")

# Команда /start
@dp.message(Command("start"))
async def start(message: types.Message):
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Выгрузить посты за текущий месяц")],
            [KeyboardButton(text="Выгрузить всю таблицу")],
            [KeyboardButton(text="Выгрузить посты за определённый месяц")]
        ],
        resize_keyboard=True
    )
    await message.answer("Привет! Выберите действие:", reply_markup=keyboard)

# Обработка нажатия кнопки "Выгрузить посты за текущий месяц"
@dp.message(lambda message: message.text == "Выгрузить посты за текущий месяц")
async def send_monthly_posts(message: types.Message):
    current_month = datetime.now().month
    monthly_file = create_monthly_excel(current_month)
    input_file = FSInputFile(monthly_file)  # Используем FSInputFile для файла из файловой системы
    await message.answer_document(input_file, caption=f"Посты за {datetime.now().year}-{current_month:02d}")
    os.remove(monthly_file)  # Удаляем временный файл

# Обработка нажатия кнопки "Выгрузить всю таблицу"
@dp.message(lambda message: message.text == "Выгрузить всю таблицу")
async def send_full_table(message: types.Message):
    full_file = create_full_excel()
    input_file = FSInputFile(full_file)  # Используем FSInputFile для файла из файловой системы
    await message.answer_document(input_file, caption="Вся таблица с постами")
    os.remove(full_file)  # Удаляем временный файл

# Обработка нажатия кнопки "Выгрузить посты за определённый месяц"
@dp.message(lambda message: message.text == "Выгрузить посты за определённый месяц")
async def choose_month(message: types.Message):
    # Создаем инлайн-клавиатуру для выбора месяца
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Январь", callback_data="month_1")],
        [InlineKeyboardButton(text="Февраль", callback_data="month_2")],
        [InlineKeyboardButton(text="Март", callback_data="month_3")],
        [InlineKeyboardButton(text="Апрель", callback_data="month_4")],
        [InlineKeyboardButton(text="Май", callback_data="month_5")],
        [InlineKeyboardButton(text="Июнь", callback_data="month_6")],
        [InlineKeyboardButton(text="Июль", callback_data="month_7")],
        [InlineKeyboardButton(text="Август", callback_data="month_8")],
        [InlineKeyboardButton(text="Сентябрь", callback_data="month_9")],
        [InlineKeyboardButton(text="Октябрь", callback_data="month_10")],
        [InlineKeyboardButton(text="Ноябрь", callback_data="month_11")],
        [InlineKeyboardButton(text="Декабрь", callback_data="month_12")],
    ])
    await message.answer("Выберите месяц текущего года:", reply_markup=keyboard)

# Обработка выбора месяца
@dp.callback_query(lambda query: query.data.startswith("month_"))
async def handle_month_selection(query: types.CallbackQuery):
    month = int(query.data.split("_")[1])  # Получаем номер месяца из callback_data
    monthly_file = create_monthly_excel(month)
    input_file = FSInputFile(monthly_file)  # Используем FSInputFile для файла из файловой системы
    await query.message.answer_document(input_file, caption=f"Посты за {datetime.now().year}-{month:02d}")
    os.remove(monthly_file)  # Удаляем временный файл
    await query.answer()  # Закрываем инлайн-клавиатуру

# Функция для фильтрации постов за текущий месяц
def filter_posts_by_current_month():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    current_month = datetime.now().month
    current_year = datetime.now().year

    filtered_posts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            # Пробуем парсить дату
            post_date = datetime.strptime(row[0], '%Y-%m-%d')
        except ValueError:
            # Если формат даты неизвестен, пропускаем запись
            continue

        if post_date.month == current_month and post_date.year == current_year:
            filtered_posts.append(row)

    return filtered_posts

# Функция для фильтрации постов за определённый месяц текущего года
def filter_posts_by_month(month):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    current_year = datetime.now().year

    filtered_posts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            # Пробуем парсить дату
            post_date = datetime.strptime(row[0], '%Y-%m-%d')
        except ValueError:
            # Если формат даты неизвестен, пропускаем запись
            continue

        if post_date.month == month and post_date.year == current_year:
            filtered_posts.append(row)

    return filtered_posts

# Функция для создания нового Excel-файла с постами за определённый месяц текущего года
def create_monthly_excel(month):
    current_year = datetime.now().year
    filtered_posts = filter_posts_by_month(month)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Posts"
    sheet.append(["Дата публикации", "Название канала", "Содержание поста", "Количество реакций"])
    for post in filtered_posts:
        sheet.append(post)
    monthly_file = f"posts_{current_year}-{month:02d}.xlsx"
    workbook.save(monthly_file)
    return monthly_file

# Функция для выгрузки всей таблицы
def create_full_excel():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    full_file = "all_posts.xlsx"
    workbook.save(full_file)
    return full_file

# Запуск бота и клиента Telethon
async def main():
    # Запуск aiogram бота
    bot_task = asyncio.create_task(dp.start_polling(bot))

    # Запуск Telethon клиента
    phone_number = input("Введите ваш номер телефона (в формате +79998887766): ")
    await client.start(phone_number)
    print("Бот запущен и отслеживает каналы...")
    client_task = asyncio.create_task(client.run_until_disconnected())

    # Ожидание завершения задач
    await asyncio.gather(bot_task, client_task)

if __name__ == '__main__':
    asyncio.run(main())